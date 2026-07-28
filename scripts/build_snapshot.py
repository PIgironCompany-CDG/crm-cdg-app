#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_snapshot.py — Automazione (GitHub Actions).
Legge il database Excel dalla App Folder di Dropbox, costruisce snapshot.json
(la vista che l'app browser mostra) e lo ricarica nella stessa App Folder.

NON scrive nulla nel repository: i dati restano su Dropbox.

Variabili d'ambiente (dai GitHub Secrets):
  DROPBOX_APP_KEY        app key dell'app Dropbox (pubblica)
  DROPBOX_REFRESH_TOKEN  refresh token generato una tantum (segreto)

Percorsi nella App Folder (relativi alla radice dell'app):
  /crm-database.xlsx   database di record
  /snapshot.json       vista generata per l'app
"""
import os, io, json, datetime, sys
import requests
import openpyxl

APP_KEY = os.environ["DROPBOX_APP_KEY"]
REFRESH_TOKEN = os.environ["DROPBOX_REFRESH_TOKEN"]
DB_PATH = "/crm-database.xlsx"
SNAPSHOT_PATH = "/snapshot.json"
OPERATIVA_PATH = "/operativa.json"   # dati esposizione/magazzino consolidati da ingest.py

TOKEN_URL = "https://api.dropbox.com/oauth2/token"
DOWNLOAD_URL = "https://content.dropboxapi.com/2/files/download"
UPLOAD_URL = "https://content.dropboxapi.com/2/files/upload"

# mappa: chiave_json -> intestazione colonna nel foglio Anagrafica_Master
FIELDS = {
 "codice":"Codice","ragione":"Ragione sociale","piva":"P.IVA","settore":"Settore",
 "tipologia":"Tipologia","status":"Status","priorita":"Priorità","stato_attivita":"Stato attività",
 "regione":"Regione","provincia":"Provincia","comune":"Comune","indirizzo":"Indirizzo",
 "email":"Email","telefono":"Telefono","referente":"Referente acquisti","proprieta":"Proprietà",
 "owner":"Owner","note":"Note","anno_bilancio":"Anno bilancio","fatturato_bilancio":"Fatturato bilancio (€)",
 "utile":"Utile/Perdita (€)","dipendenti":"N. dipendenti","rating":"Rating credito",
 "punteggio":"Punteggio credito","limite_credito":"Limite credito report (€)",
 "gg_silenzio":"Giorni di silenzio","fatturato_storico":"Fatturato storico (€)","n_ordini":"N. ordini",
}

def get_access_token():
    r = requests.post(TOKEN_URL, data={
        "grant_type": "refresh_token",
        "refresh_token": REFRESH_TOKEN,
        "client_id": APP_KEY,
    }, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]

def download(token, path):
    r = requests.post(DOWNLOAD_URL, headers={
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": json.dumps({"path": path}),
    }, timeout=60)
    r.raise_for_status()
    return r.content

def download_opt(token, path):
    """Download che ritorna None se il file non esiste (no eccezione)."""
    r = requests.post(DOWNLOAD_URL, headers={
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": json.dumps({"path": path}),
    }, timeout=60)
    return r.content if r.status_code == 200 else None

def upload(token, path, data: bytes):
    r = requests.post(UPLOAD_URL, headers={
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": json.dumps({"path": path, "mode": "overwrite", "mute": True}),
        "Content-Type": "application/octet-stream",
    }, data=data, timeout=60)
    r.raise_for_status()
    return r.json()

def conv(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v

def build(xlsx_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    M = wb["Anagrafica_Master"]; H = [c.value for c in M[1]]; HX = {h:i for i,h in enumerate(H) if h}

    # Registro operativo (foglio Pipeline righe 5-123) = fonte affidabile dei follow-up
    reg = {}
    try:
        P = wb["Pipeline"]
        for row in P.iter_rows(min_row=5, max_row=123, values_only=True):
            if row and row[0]:
                reg[str(row[0]).strip()] = {
                    "ultimo": row[6], "esito": row[8], "prossima": row[9],
                    "followup": row[10], "ntent": row[11], "stato": row[12],
                }
    except Exception:
        pass

    # ---- Vendite: KPI YTD (pari periodo) + serie mensile + per-cliente ----
    import collections as _c
    vend = {}
    mensili = _c.OrderedDict((f"{m:02d}", {"fatturato":0.0,"tonnellate":0.0}) for m in range(1,13))
    ytd = {2025:{"fatt":0.0,"ton":0.0,"rows":0,"cli":set()}, 2026:{"fatt":0.0,"ton":0.0,"rows":0,"cli":set()}}
    ref_doy = datetime.date.today().timetuple().tm_yday
    grid = {}   # cruscotto in testa al foglio Vendite (righe 1-28)
    try:
        V = wb["Vendite"]; VX = {}
        for r, row in enumerate(V.iter_rows(min_row=1, values_only=True), 1):
            if r <= 28:
                for c, val in enumerate(row[:14], 1):
                    if val is not None: grid[(r,c)] = val
            elif r == 31:
                VX = {str(x).strip(): i for i, x in enumerate(row) if x}
            elif r >= 32 and VX:
                cod = row[VX["Cod.Cli"]] if "Cod.Cli" in VX else None
                if not cod: continue
                cod=str(cod).strip()
                dt = row[VX.get("Data")] if "Data" in VX else None
                if not isinstance(dt,(datetime.datetime,datetime.date)): continue
                imp = row[VX.get("Importo")] or 0
                q   = (row[VX.get("Quantità (kg)")] or 0)/1000
                if dt.year==2026:
                    d = vend.setdefault(cod, {"fatt":0.0,"ton":0.0}); d["fatt"]+=imp; d["ton"]+=q
                    mensili[f"{dt.month:02d}"]["fatturato"]+=imp; mensili[f"{dt.month:02d}"]["tonnellate"]+=q
                if dt.year in ytd and dt.timetuple().tm_yday<=ref_doy:
                    y=ytd[dt.year]; y["fatt"]+=imp; y["ton"]+=q; y["rows"]+=1; y["cli"].add(cod)
    except Exception:
        pass
    def _et(y): return round(y["fatt"]/y["ton"]) if y["ton"] else 0
    totali = {
        "fatturato": round(ytd[2026]["fatt"]), "tonnellate": round(ytd[2026]["ton"],1),
        "ordini": ytd[2026]["rows"], "clienti_attivi": len(ytd[2026]["cli"]), "et_medio": _et(ytd[2026]),
        "fatturato_2025": round(ytd[2025]["fatt"]), "tonnellate_2025": round(ytd[2025]["ton"],1), "et_2025": _et(ytd[2025]),
    }
    # ---- Vista operativa: esposizione finanziaria + magazzino (cruscotto foglio Vendite) ----
    def g(r,c): return grid.get((r,c))
    def gnum(r,c):
        v=g(r,c)
        try: return round(float(v),2)
        except: return None
    operativa = {}
    try:
        operativa["esposizione"] = {
            "clienti": gnum(8,1), "merce_a_terra": gnum(8,6), "totale": gnum(8,11),
            "voci": [{"voce":g(r,9), "eur":gnum(r,12), "usd":gnum(r,14)} for r in range(20,24) if g(r,9)],
        }
        operativa["magazzino"] = {
            "disponibile_ton": gnum(15,1), "uscite_ytd_ton": gnum(15,6), "mesi_copertura": gnum(15,11),
        }
        lotti=[]; r=20
        while g(r,1) and str(g(r,1)).strip().upper()!="TOTALE" and r<30:
            lotti.append({"qualita":g(r,1),"lotto":g(r,2),"ton":gnum(r,3),"prezzo_usd":gnum(r,4),"prezzo_eur":gnum(r,5),"note":g(r,6)})
            r+=1
        operativa["lotti"]=lotti
        mk=["01","02","03","04","05","06","07","08","09","10","11","12"]
        operativa["uscite_mensili"]=[{"mese":mk[i],"tonnellate":gnum(28,i+1) or 0} for i in range(12)]
    except Exception:
        operativa={}

    clients = []
    for row in M.iter_rows(min_row=2, values_only=True):
        cod = row[HX["Codice"]] if "Codice" in HX else None
        if not cod:
            continue
        rec = {}
        for k, col in FIELDS.items():
            i = HX.get(col)
            rec[k] = conv(row[i]) if (i is not None and i < len(row)) else None
        rg = reg.get(str(cod).strip())
        if rg:
            rec["esito"] = conv(rg.get("esito"))
            rec["prossima_azione"] = conv(rg.get("prossima"))
            rec["data_ultima_azione"] = conv(rg.get("ultimo"))
            rec["data_followup"] = conv(rg.get("followup"))
            rec["n_tentativi"] = rg.get("ntent")
            rec["stato_followup"] = conv(rg.get("stato"))
        va = vend.get(str(cod).strip())
        rec["fatturato_2026"]  = round(va["fatt"]) if va else 0
        rec["tonnellate_2026"] = round(va["ton"],1) if va else 0
        clients.append(rec)
    wb.close()
    return {
        "generato": datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "n_clienti": len(clients),
        "clienti": clients,
        "vendite": {
            "mensili": [{"mese":k, "fatturato":round(v["fatturato"]), "tonnellate":round(v["tonnellate"],1)} for k,v in mensili.items()],
            "totali": totali,
        },
        "operativa": operativa,
    }

def main():
    token = get_access_token()
    print("Token ottenuto. Scarico il database...")
    xlsx = download(token, DB_PATH)
    print(f"Database scaricato ({len(xlsx)} byte). Costruisco lo snapshot...")
    snap = build(xlsx)
    # Operativa CENTRALIZZATA: /operativa.json (prodotto da ingest.py dai file esposizione/magazzino).
    opraw = download_opt(token, OPERATIVA_PATH)
    try:
        snap["operativa"] = json.loads(opraw) if opraw else {}
    except Exception:
        snap["operativa"] = {}
    op = snap["operativa"]
    print(f"OPERATIVA: righe_esposizione={len((op.get('esposizione') or {}).get('righe',[]))} "
          f"giacenze={len(op.get('giacenze',[]))} "
          f"uscite_valorizzate={sum(1 for m in op.get('uscite_mensili',[]) if m.get('tonnellate'))} "
          f"aggiornato={op.get('aggiornato')}")
    if not op.get('esposizione') and not op.get('giacenze'):
        print("NOTA: operativa.json assente o vuoto. Carica i file in /caricamenti; l'ingest gira alle 10:00 e 18:00.")
    blob = json.dumps(snap, ensure_ascii=False).encode("utf-8")
    upload(token, SNAPSHOT_PATH, blob)
    print(f"Snapshot pubblicato: {snap['n_clienti']} clienti, {len(blob)} byte.")

if __name__ == "__main__":
    try:
        sys.exit(main())
    except requests.HTTPError as e:
        print("Errore HTTP Dropbox:", e.response.status_code, e.response.text[:500])
        sys.exit(1)
